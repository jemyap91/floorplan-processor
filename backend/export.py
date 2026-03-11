"""Excel export for floorplan processing results."""
import io
import math
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from backend.models.room import ProjectData, RoomData


def build_excel(project: ProjectData, rooms: list[RoomData]) -> bytes:
    """Build a rich Excel workbook with all structured floorplan data.

    Returns the workbook as bytes (xlsx format).
    """
    wb = Workbook()

    _build_summary_sheet(wb, project, rooms)
    _build_room_schedule_sheet(wb, project, rooms)
    _build_wall_segments_sheet(wb, project, rooms)
    _build_polygon_sheet(wb, project, rooms)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D4D4D4"),
    right=Side(style="thin", color="D4D4D4"),
    top=Side(style="thin", color="D4D4D4"),
    bottom=Side(style="thin", color="D4D4D4"),
)
_NUM_FMT_2DP = "0.00"
_NUM_FMT_1DP = "0.0"
_NUM_FMT_INT = "#,##0"
_LABEL_FONT = Font(bold=True, color="374151")
_LABEL_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")


def _style_header_row(ws, num_cols: int):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.freeze_panes = "A2"


def _auto_width(ws, min_width: int = 8, max_width: int = 30):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            lengths.append(len(val))
        best = min(max(max(lengths, default=min_width), min_width), max_width)
        ws.column_dimensions[col_letter].width = best + 2


def _rgb_to_hex(r: int, g: int, b: int) -> str:
    """Convert RGB to hex string for openpyxl (no # prefix)."""
    return f"{r:02X}{g:02X}{b:02X}"


# ---------------------------------------------------------------------------
# Sheet 1: Summary
# ---------------------------------------------------------------------------


def _build_summary_sheet(wb: Workbook, project: ProjectData, rooms: list[RoomData]):
    ws = wb.active
    ws.title = "Summary"

    total_area_sqm = sum(r.area_sqm or 0 for r in rooms)
    total_area_px = sum(r.area_px for r in rooms)
    total_perim_m = sum(r.perimeter_m or 0 for r in rooms)

    rows = [
        ("Project Name", project.name or "Untitled"),
        ("Date Processed", project.created_at.strftime("%Y-%m-%d %H:%M") if project.created_at else "—"),
        ("", ""),
        ("Scale (px/m)", f"{project.scale_px_per_meter:.2f}" if project.scale_px_per_meter else "Not set"),
        ("Scale (m/px)", f"{1/project.scale_px_per_meter:.5f}" if project.scale_px_per_meter else "—"),
        ("Scale Source", project.scale_source),
        ("", ""),
        ("Total Rooms", len(rooms)),
        ("Total Area (m²)", round(total_area_sqm, 2) if total_area_sqm else "—"),
        ("Total Area (px)", round(total_area_px, 0)),
        ("Total Perimeter (m)", round(total_perim_m, 2) if total_perim_m else "—"),
        ("", ""),
        ("Room Types", ""),
    ]

    # Count room types
    type_counts: dict[str, int] = {}
    for r in rooms:
        t = r.room_type or "unknown"
        type_counts[t] = type_counts.get(t, 0) + 1
    for t, count in sorted(type_counts.items(), key=lambda x: -x[1]):
        rows.append((f"  {t}", count))

    for row_idx, (label, value) in enumerate(rows, start=1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = _LABEL_FONT
        if label:
            label_cell.fill = _LABEL_FILL
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30


# ---------------------------------------------------------------------------
# Sheet 2: Room Schedule
# ---------------------------------------------------------------------------

_ROOM_HEADERS = [
    "#",
    "Room Name",
    "Room Type",
    "Fill Color",
    "Area (m²)",
    "Area (px)",
    "Perimeter (m)",
    "Perimeter (px)",
    "No. of Walls",
    "Bbox Width (m)",
    "Bbox Height (m)",
    "Bbox Width (px)",
    "Bbox Height (px)",
    "Aspect Ratio",
    "Compactness",
    "Centroid X",
    "Centroid Y",
    "Source",
    "Confidence",
]


def _compute_room_metrics(room: RoomData, px_per_m: float | None) -> dict:
    """Compute derived metrics from a room's polygon."""
    poly = room.boundary_polygon
    n_walls = max(len(room.boundary_lengths_px), len(poly) - 1 if len(poly) > 1 else 0)

    # Bounding box
    if poly:
        xs = [p[0] for p in poly]
        ys = [p[1] for p in poly]
        bbox_w_px = max(xs) - min(xs)
        bbox_h_px = max(ys) - min(ys)
    else:
        bbox_w_px = bbox_h_px = 0

    bbox_w_m = bbox_w_px / px_per_m if px_per_m else None
    bbox_h_m = bbox_h_px / px_per_m if px_per_m else None

    # Aspect ratio (0–1, 1 = square)
    long_side = max(bbox_w_px, bbox_h_px, 1)
    short_side = min(bbox_w_px, bbox_h_px)
    aspect = short_side / long_side

    # Compactness (isoperimetric quotient): 4πA/P²
    perim = room.perimeter_px or 1
    compactness = (4 * math.pi * room.area_px) / (perim ** 2) if perim > 0 else 0

    return {
        "n_walls": n_walls,
        "bbox_w_px": bbox_w_px,
        "bbox_h_px": bbox_h_px,
        "bbox_w_m": bbox_w_m,
        "bbox_h_m": bbox_h_m,
        "aspect": round(aspect, 3),
        "compactness": round(compactness, 3),
    }


def _build_room_schedule_sheet(
    wb: Workbook, project: ProjectData, rooms: list[RoomData]
):
    ws = wb.create_sheet("Room Schedule")
    px_per_m = project.scale_px_per_meter

    # Headers
    for col, header in enumerate(_ROOM_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(_ROOM_HEADERS))

    for row_idx, room in enumerate(rooms, start=2):
        m = _compute_room_metrics(room, px_per_m)

        # Determine fill color hex from room's fill_color_rgb if available
        fill_hex = ""
        fill_color = getattr(room, "fill_color_rgb", None)
        if fill_color and len(fill_color) == 3:
            r, g, b = int(fill_color[0]), int(fill_color[1]), int(fill_color[2])
            fill_hex = f"#{r:02X}{g:02X}{b:02X}"

        values = [
            row_idx - 1,                        # #
            room.name,                           # Room Name
            room.room_type,                      # Room Type
            fill_hex,                            # Fill Color
            round(room.area_sqm, 2) if room.area_sqm else None,
            round(room.area_px, 0),
            round(room.perimeter_m, 2) if room.perimeter_m else None,
            round(room.perimeter_px, 0),
            m["n_walls"],
            round(m["bbox_w_m"], 2) if m["bbox_w_m"] else None,
            round(m["bbox_h_m"], 2) if m["bbox_h_m"] else None,
            round(m["bbox_w_px"], 0),
            round(m["bbox_h_px"], 0),
            m["aspect"],
            m["compactness"],
            round(room.centroid[0], 1),
            round(room.centroid[1], 1),
            room.source,
            round(room.confidence, 2),
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = _THIN_BORDER

        # Apply fill color to the Fill Color cell
        if fill_hex:
            color_cell = ws.cell(row=row_idx, column=4)
            hex_code = _rgb_to_hex(r, g, b)
            color_cell.fill = PatternFill(
                start_color=hex_code, end_color=hex_code, fill_type="solid"
            )
            # Use dark text on light fills, light text on dark fills
            brightness = (r * 299 + g * 587 + b * 114) / 1000
            text_color = "000000" if brightness > 128 else "FFFFFF"
            color_cell.font = Font(color=text_color, size=10)

    _auto_width(ws)
    # Widen the name column
    ws.column_dimensions["B"].width = 25


# ---------------------------------------------------------------------------
# Sheet 3: Wall Segments
# ---------------------------------------------------------------------------

_WALL_HEADERS = [
    "Room #",
    "Room Name",
    "Wall #",
    "Length (m)",
    "Length (px)",
    "Start X (px)",
    "Start Y (px)",
    "End X (px)",
    "End Y (px)",
    "Orientation",
]


def _wall_orientation(x1: float, y1: float, x2: float, y2: float) -> str:
    """Classify wall segment as Horizontal, Vertical, or Diagonal."""
    dx = abs(x2 - x1)
    dy = abs(y2 - y1)
    if dx < 1 and dy < 1:
        return "Point"
    angle = math.degrees(math.atan2(dy, dx))
    if angle < 15:
        return "Horizontal"
    if angle > 75:
        return "Vertical"
    return f"Diagonal ({angle:.0f}°)"


def _build_wall_segments_sheet(
    wb: Workbook, project: ProjectData, rooms: list[RoomData]
):
    ws = wb.create_sheet("Wall Segments")
    px_per_m = project.scale_px_per_meter

    for col, header in enumerate(_WALL_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(_WALL_HEADERS))

    row_idx = 2
    for room_idx, room in enumerate(rooms):
        poly = room.boundary_polygon
        lengths_px = room.boundary_lengths_px
        lengths_m = room.boundary_lengths_m

        n_segments = len(lengths_px) if lengths_px else (len(poly) - 1 if len(poly) > 1 else 0)

        for seg_idx in range(n_segments):
            # Vertex coordinates
            if seg_idx < len(poly) - 1:
                x1, y1 = poly[seg_idx]
                x2, y2 = poly[seg_idx + 1]
            elif seg_idx < len(poly):
                x1, y1 = poly[seg_idx]
                x2, y2 = poly[0]
            else:
                x1 = y1 = x2 = y2 = 0

            length_px = lengths_px[seg_idx] if seg_idx < len(lengths_px) else None
            length_m = lengths_m[seg_idx] if lengths_m and seg_idx < len(lengths_m) else None

            orientation = _wall_orientation(x1, y1, x2, y2)

            values = [
                room_idx + 1,
                room.name,
                seg_idx + 1,
                round(length_m, 2) if length_m is not None else None,
                round(length_px, 1) if length_px is not None else None,
                round(x1, 1),
                round(y1, 1),
                round(x2, 1),
                round(y2, 1),
                orientation,
            ]

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = _THIN_BORDER

            row_idx += 1

    _auto_width(ws)
    ws.column_dimensions["B"].width = 25


# ---------------------------------------------------------------------------
# Sheet 4: Polygon Coordinates
# ---------------------------------------------------------------------------

_POLY_HEADERS = [
    "Room #",
    "Room Name",
    "Vertex #",
    "X (px)",
    "Y (px)",
    "X (m)",
    "Y (m)",
]


def _build_polygon_sheet(
    wb: Workbook, project: ProjectData, rooms: list[RoomData]
):
    ws = wb.create_sheet("Polygon Coordinates")
    px_per_m = project.scale_px_per_meter

    for col, header in enumerate(_POLY_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, len(_POLY_HEADERS))

    row_idx = 2
    for room_idx, room in enumerate(rooms):
        for vert_idx, point in enumerate(room.boundary_polygon):
            x_px, y_px = point[0], point[1]
            x_m = x_px / px_per_m if px_per_m else None
            y_m = y_px / px_per_m if px_per_m else None

            values = [
                room_idx + 1,
                room.name,
                vert_idx + 1,
                round(x_px, 1),
                round(y_px, 1),
                round(x_m, 4) if x_m is not None else None,
                round(y_m, 4) if y_m is not None else None,
            ]

            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = _THIN_BORDER

            row_idx += 1

    _auto_width(ws)
    ws.column_dimensions["B"].width = 25
